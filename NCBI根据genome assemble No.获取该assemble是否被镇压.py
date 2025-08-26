import asyncio
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import openpyxl
import logging

async def check_genome_id(genome_id):
    chrome_options = Options()
    chrome_options.add_argument("--headless")  # Run the browser in headless mode (without GUI)
    driver = webdriver.Chrome(options=chrome_options)

    url = f"https://www.ncbi.nlm.nih.gov/datasets/genome/{genome_id}/"
    driver.get(url)
    html = driver.page_source

    if "is suppressed" in html:
        return True
    else:
        return False

    driver.quit()

async def process_excel(input_file, output_file):
    # Load the HH2.xlsx file
    logging.info("Loading HH2.xlsx file...")
    wb_original = openpyxl.load_workbook(input_file)
    sheet_original = wb_original.active
    logging.info(f"Loaded {sheet_original.title} sheet with {sheet_original.max_row} rows")

    # Create a new Excel file
    logging.info("Creating new Excel file...")
    wb_new = openpyxl.Workbook()
    sheet_new = wb_new.active
    logging.info("New Excel file created")

    # Set the header row
    logging.info("Setting header row...")
    sheet_new.cell(row=1, column=1, value='Genome ID')
    sheet_new.cell(row=1, column=2, value='Is Suppressed')
    logging.info("Header row set")

    # Iterate through the first column of the original sheet
    row_num = 2
    for row in sheet_original.iter_rows(min_row=1, max_col=1, values_only=False):
        cell = row[0]
        genome_id = cell.value.strip()  # Get the genome ID from the cell value
        logging.info(f"Processing genome ID: {genome_id}")

        is_suppressed = await check_genome_id(genome_id)
        logging.info(f"Result for genome ID {genome_id}: {is_suppressed}")

        sheet_new.cell(row=row_num, column=1, value=genome_id)
        sheet_new.cell(row=row_num, column=2, value=is_suppressed)
        row_num += 1

    # Save the new Excel file
    logging.info("Saving new Excel file...")
    wb_new.save(output_file)
    logging.info("New Excel file saved")

if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    asyncio.run(process_excel('HH2.xlsx', 'HH2_results.xlsx'))