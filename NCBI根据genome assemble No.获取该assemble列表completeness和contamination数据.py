import asyncio
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import openpyxl
import logging
import re

async def get_completeness_contamination(genome_id):
    chrome_options = Options()
    chrome_options.add_argument("--headless")  # Run the browser in headless mode (without GUI)
    driver = webdriver.Chrome(options=chrome_options)

    url = f"https://www.ncbi.nlm.nih.gov/datasets/genome/{genome_id}/"
    driver.get(url)
    html = driver.page_source

    completeness_pattern = r"Completeness:\s*(\d+(?:\.\d+)?)\s*%"
    contamination_pattern = r"Contamination:\s*(\d+(?:\.\d+)?)\s*%"

    completeness_match = re.search(completeness_pattern, html)
    contamination_match = re.search(contamination_pattern, html)

    if completeness_match and contamination_match:
        completeness = float(completeness_match.group(1))
        contamination = float(contamination_match.group(1))
        return completeness, contamination
    else:
        return None, None

    driver.quit()

async def process_excel(input_file, output_file):
    # Load the HH2.xlsx file
    logging.info("Loading HH2.xlsx file...")
    wb_original = openpyxl.load_workbook(input_file)
    sheet_original = wb_original.active
    logging.info(f"Loaded {sheet_original.title} sheet with {sheet_original.max_row} rows")

    # Create a new Excel file
    logging.info("Creating new Excel file...")
    wb_new = openpyxl.Workbook()
    sheet_new = wb_new.active
    logging.info("New Excel file created")

    # Set the header row
    logging.info("Setting header row...")
    sheet_new.cell(row=1, column=1, value='Genome ID')
    sheet_new.cell(row=1, column=2, value='Completeness')
    sheet_new.cell(row=1, column=3, value='Contamination')
    logging.info("Header row set")

    # Iterate through the first column of the original sheet
    row_num = 2
    for row in sheet_original.iter_rows(min_row=2, max_col=1, values_only=True):
        genome_id = row[0].strip()  # Get the genome ID from the cell value
        logging.info(f"Processing genome ID: {genome_id}")

        completeness, contamination = await get_completeness_contamination(genome_id)
        logging.info(f"Result for genome ID {genome_id}: Completeness={completeness}, Contamination={contamination}")

        sheet_new.cell(row=row_num, column=1, value=genome_id)
        if completeness is not None:
            sheet_new.cell(row=row_num, column=2, value=completeness)
        if contamination is not None:
            sheet_new.cell(row=row_num, column=3, value=contamination)
        row_num += 1

    # Save the new Excel file
    logging.info("Saving new Excel file...")
    wb_new.save(output_file)
    logging.info("New Excel file saved")

if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    asyncio.run(process_excel('HH2.xlsx', 'HH2_results_C_C102.xlsx'))